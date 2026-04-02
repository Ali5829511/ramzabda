import pandas as pd

path = r'C:\Users\aliay\Downloads\units.report.2026-03-19  الواحدات.xlsx'

# Peek structure
import openpyxl
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'[{name}] rows={ws.max_row} cols={ws.max_column}')
wb.close()

# Read all rows without header first
df_raw = pd.read_excel(path, sheet_name=0, header=None, nrows=5)
print('\nFirst 5 rows:')
print(df_raw.to_string())
